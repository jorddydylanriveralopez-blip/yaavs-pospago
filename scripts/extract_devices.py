#!/usr/bin/env python3
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception as e:
    raise RuntimeError('openpyxl is required. Run: pip3 install openpyxl')

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / '_refs' / 'PLANTILLA PRECIOS 1 sept.xlsx'
OUT = ROOT / 'assets' / 'data' / 'devices.json'
OUT.parent.mkdir(parents=True, exist_ok=True)

if not XLSX.exists():
    raise SystemExit(f'Excel not found: {XLSX}')

wb = load_workbook(XLSX, data_only=True)
# Assume data is in the first sheet
ws = wb[wb.sheetnames[0]]

headers = []
rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        headers = [str(c).strip() if c is not None else '' for c in row]
        continue
    if all(c is None for c in row):
        continue
    item = {}
    for h, c in zip(headers, row):
        if h:
            item[h] = c
    rows.append(item)

# Normalize keys: lowercase and strip
devices = []
for r in rows:
    # Try to find common fields
    name = None
    price = None
    for k, v in r.items():
        lk = k.lower()
        if 'nombre' in lk or 'modelo' in lk or 'device' in lk or 'equipo' in lk or 'modelo' in lk:
            if v:
                name = str(v).strip()
        if 'precio' in lk or 'price' in lk or 'pvp' in lk or 'precio final' in lk:
            price = v
    # fallback: first non-empty string cell
    if not name:
        for v in r.values():
            if isinstance(v, str) and v.strip():
                name = v.strip()
                break
    devices.append({
        'raw': r,
        'name': name,
        'price': price,
    })

with OUT.open('w', encoding='utf-8') as f:
    json.dump(devices, f, ensure_ascii=False, indent=2)

print('Wrote', OUT)
